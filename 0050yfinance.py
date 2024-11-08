import pandas as pd
from openpyxl.reader.excel import load_workbook
import yfinance as yf
from datetime import datetime, timedelta
from plot_stock import stock_trend

def read_stock_codes(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    stock_codes = []

    # for迴圈取得50檔股票
    for row in range(19, 69):
        stock_code = ws['A'+str(row)].value
        stock_codes.append(str(stock_code) + '.TW')
    return stock_codes


#定義布林通道的計算
def bollinger_band(historical_data, window=20, std_dev=2):
    # 檢查是否有足夠數據計算
    if historical_data.shape[0] < window:
        return {'Last Close Price': None, 'Lower Band': None, 'Buy Signal': False}

    #計算均線和標準差
    historical_data['SMA'] = historical_data['Close'].rolling(window=window).mean()
    historical_data['Std Dev'] = historical_data['Close'].rolling(window=window).std()
    historical_data['Upper Band'] = historical_data['SMA'] + (historical_data['Std Dev']*std_dev)
    historical_data['Lower Band'] = historical_data['SMA'] - (historical_data['Std Dev'] * std_dev)

    #取最後收盤價和下軌的值
    last_close = historical_data['Close'].iloc[-1].values
    lower_band = historical_data['Lower Band'].iloc[-1]

    
    #判斷是否買入
    if last_close < lower_band:
        buy_signal = True
    else:
        buy_signal = False

    return{'Last Close Price': last_close,
           'Lower Band': lower_band,
           'Buy Signal': buy_signal
    }

if __name__ == '__main__':
    #提取股票代號
    stock_info = read_stock_codes('0050.xlsx')

    # 計算時間範圍
    end_date = datetime.now()
    start_date = end_date - pd.DateOffset(months=6)

    # 初始化空的DateFrame用於儲存historical_data
    buy_signal_df = pd.DataFrame()
    buy_signals = []

    #主迴圈
    for code in stock_info:
        historical_data = yf.download(code, start=start_date.strftime('%Y-%m-%d'),
                                  end=end_date.strftime('%Y-%m-%d'))
        if historical_data.empty:
            print(f"None for code {code}")
            continue

        # 添加一列標記股票代號並去掉TW
        historical_data['code'] = code[:-3]

        # rest index並去除時區
        historical_data.reset_index(inplace=True)
        historical_data['Date'] = historical_data['Date'].dt.tz_localize(None)

        # 計算布林通道
        bol_result = bollinger_band(historical_data)

        # 將適合買入的股票加入列表
        if bol_result['Buy Signal']:
            buy_signals.append({'Stock Code': code,
                            'Last Close Price': bol_result['Last Close Price'],
                            'Lower Band': bol_result['Lower Band']})
            buy_signal_df = pd.concat([buy_signal_df, historical_data], ignore_index=True)

            #畫出布林通道
            stock_trend(historical_data, code)


    if not buy_signal_df.empty:
        # 處理multiindex，將其轉成flatten index
        if isinstance(buy_signal_df.columns, pd.MultiIndex):
            buy_signal_df.columns = [' '.join(col).strip() for col in buy_signal_df.columns.values]


    #數據存入excel檔
    buy_signal_df.to_excel('buy_signal stocks.xlsx', index=False)

    print("符合買入條件的股票有:")
    print(buy_signals)
